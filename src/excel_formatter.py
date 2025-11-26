"""
Excel Export with Conditional Formatting Module

This module handles exporting race data to Excel with color highlighting
applied only to rows corresponding to bet-worthy races.

Uses openpyxl for fine-grained control over cell formatting.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


# ===== COLOR CONFIGURATION =====
# Customize these colors for bet-worthy race highlighting

# Position-based colors for bet-worthy races
FIRST_PLACE_COLOR = "90EE90"   # Light green (top pick)
SECOND_PLACE_COLOR = "FFD580"  # Light orange (second pick)
THIRD_PLACE_COLOR = "FFB6C1"   # Light red/pink (third pick)
OTHER_BET_WORTHY_COLOR = "FFFF99"  # Light yellow (other dogs in bet-worthy races)

# Font color for bet-worthy races (keep black for readability)
BET_WORTHY_FONT_COLOR = "000000"  # Black

# Header background color
HEADER_FILL_COLOR = "4472C4"  # Blue

# Header font color
HEADER_FONT_COLOR = "FFFFFF"  # White


def export_to_excel_with_formatting(df, bet_worthy_races, output_path):
    """
    Export DataFrame to Excel with conditional formatting for bet-worthy races.
    
    For bet-worthy races, dogs are colored by their position (based on FinalScore):
    - 1st place (top pick): Light green
    - 2nd place: Light orange
    - 3rd place: Light red/pink
    - Other dogs in bet-worthy race: Light yellow
    
    Args:
        df: DataFrame with race data (must include FinalScore column)
        bet_worthy_races: Dict from bet_worthy.identify_bet_worthy_races()
                          Maps (Track, RaceNumber) -> {'worthy': bool, ...}
        output_path: Full path where Excel file should be saved
    """
    print(f"\n📊 Creating Excel file with conditional formatting...")
    
    # Create a new workbook and select active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Race Analysis"
    
    # Define position-based color fills
    first_place_fill = PatternFill(start_color=FIRST_PLACE_COLOR,
                                     end_color=FIRST_PLACE_COLOR,
                                     fill_type="solid")
    
    second_place_fill = PatternFill(start_color=SECOND_PLACE_COLOR,
                                      end_color=SECOND_PLACE_COLOR,
                                      fill_type="solid")
    
    third_place_fill = PatternFill(start_color=THIRD_PLACE_COLOR,
                                     end_color=THIRD_PLACE_COLOR,
                                     fill_type="solid")
    
    other_bet_worthy_fill = PatternFill(start_color=OTHER_BET_WORTHY_COLOR,
                                          end_color=OTHER_BET_WORTHY_COLOR,
                                          fill_type="solid")
    
    header_fill = PatternFill(start_color=HEADER_FILL_COLOR,
                               end_color=HEADER_FILL_COLOR,
                               fill_type="solid")
    
    header_font = Font(bold=True, color=HEADER_FONT_COLOR)
    bet_worthy_font = Font(color=BET_WORTHY_FONT_COLOR)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Pre-calculate positions for all dogs in bet-worthy races
    dog_positions = {}  # Maps (Track, RaceNumber, row_index) -> position (1, 2, 3, or None)
    
    for (track, race_num), race_info in bet_worthy_races.items():
        if race_info.get('worthy', False):
            # Get all dogs in this race
            race_dogs = df[(df['Track'] == track) & (df['RaceNumber'] == race_num)].copy()
            
            # Sort by FinalScore descending and take top 3
            if 'FinalScore' in race_dogs.columns:
                race_dogs = race_dogs.sort_values('FinalScore', ascending=False).head(3)
                
                # Assign positions (1, 2, 3) to top 3 dogs
                for position, (idx, dog) in enumerate(race_dogs.iterrows(), start=1):
                    dog_positions[(track, race_num, idx)] = position
    
    # Write headers
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Track statistics
    total_rows = 0
    highlighted_rows = 0
    position_counts = {1: 0, 2: 0, 3: 0, 'other': 0}
    
    # Write data rows
    for row_idx, (df_idx, row_data) in enumerate(df.iterrows(), start=2):
        track = row_data.get('Track', '')
        race_num = row_data.get('RaceNumber', 0)
        
        # Check if this race is bet-worthy
        race_key = (track, race_num)
        is_bet_worthy = bet_worthy_races.get(race_key, {}).get('worthy', False)
        
        # Determine position for this dog in bet-worthy race
        position = None
        if is_bet_worthy:
            position = dog_positions.get((track, race_num, df_idx), None)
            highlighted_rows += 1
            if position in [1, 2, 3]:
                position_counts[position] += 1
            else:
                position_counts['other'] += 1
        
        # Select the appropriate fill color based on position
        cell_fill = None
        if is_bet_worthy:
            if position == 1:
                cell_fill = first_place_fill
            elif position == 2:
                cell_fill = second_place_fill
            elif position == 3:
                cell_fill = third_place_fill
            else:
                cell_fill = other_bet_worthy_fill
        
        # Write each cell in the row
        for col_idx, (col_name, value) in enumerate(row_data.items(), start=1):
            # Convert lists to strings for Excel compatibility
            if isinstance(value, (list, tuple)):
                value = str(value)
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # Apply formatting only to bet-worthy races
            if is_bet_worthy and cell_fill:
                cell.fill = cell_fill
                cell.font = bet_worthy_font
        
        total_rows += 1
    
    # Auto-adjust column widths (basic approach)
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                # Skip cells that can't be converted to string
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for very long values
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Save the workbook
    wb.save(output_path)
    
    print(f"✅ Excel file saved: {output_path}")
    print(f"   Total rows: {total_rows}")
    print(f"   Highlighted (bet-worthy) rows: {highlighted_rows}")
    print(f"     - 1st place (green): {position_counts[1]}")
    print(f"     - 2nd place (orange): {position_counts[2]}")
    print(f"     - 3rd place (red): {position_counts[3]}")
    print(f"     - Other (yellow): {position_counts['other']}")
    print(f"   Non-highlighted rows: {total_rows - highlighted_rows}")
    
    return output_path


def export_simple_excel(df, output_path):
    """
    Export DataFrame to Excel without any special formatting (fallback method).
    
    Args:
        df: DataFrame to export
        output_path: Full path where Excel file should be saved
    """
    df.to_excel(output_path, index=False, engine='openpyxl')
    print(f"✅ Excel file saved (no formatting): {output_path}")
    return output_path
