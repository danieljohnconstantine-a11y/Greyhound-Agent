"""
Excel export with color-coded rankings for greyhound race predictions.
Creates Excel files with conditional formatting highlighting top 3 predictions per race.
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os


def color_code_predictions(df, output_path):
    """
    Export DataFrame to Excel with color-coded top 3 predictions per race.
    
    Args:
        df: DataFrame with race data including Track, RaceNumber, FinalScore
        output_path: Path to save the Excel file
    
    Color coding:
        - Green (#90EE90): 1st place prediction (highest FinalScore)
        - Orange (#FFA500): 2nd place prediction
        - Red (#FFB6C1): 3rd place prediction
    """
    # Sort by Track, RaceNumber, and FinalScore (descending)
    df_sorted = df.sort_values(['Track', 'RaceNumber', 'FinalScore'], 
                               ascending=[True, True, False]).reset_index(drop=True)
    
    # Save to Excel first
    df_sorted.to_excel(output_path, index=False, engine='openpyxl')
    
    # Load workbook for formatting
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Define colors
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    
    # Track which races we've processed
    current_race_key = None
    race_rank = 0
    
    # Iterate through rows (skip header)
    for idx, row in df_sorted.iterrows():
        excel_row = idx + 2  # +2 because Excel is 1-indexed and row 1 is header
        
        race_key = (row.get('Track', ''), row.get('RaceNumber', ''))
        
        # Reset rank counter when we hit a new race
        if race_key != current_race_key:
            current_race_key = race_key
            race_rank = 1
        else:
            race_rank += 1
        
        # Apply color based on rank within race
        if race_rank == 1:
            fill = green_fill
        elif race_rank == 2:
            fill = orange_fill
        elif race_rank == 3:
            fill = red_fill
        else:
            continue  # Don't color ranks beyond 3rd
        
        # Apply fill to all cells in the row
        for cell in ws[excel_row]:
            cell.fill = fill
    
    # Save the formatted workbook
    wb.save(output_path)
    print(f"✅ Color-coded Excel saved: {output_path}")
    print("   🟢 Green = 1st prediction | 🟠 Orange = 2nd prediction | 🔴 Red = 3rd prediction")


def create_color_coded_outputs(combined_df, output_dir="outputs"):
    """
    Create color-coded Excel files for all outputs.
    
    Args:
        combined_df: Combined DataFrame with all dogs
        output_dir: Directory to save output files
    """
    os.makedirs(output_dir, exist_ok=True)
    
    # 1. Full form with color-coding by race
    print("\n📊 Creating color-coded Excel outputs...")
    color_code_predictions(combined_df, 
                          os.path.join(output_dir, "todays_form_color.xlsx"))
    
    # 2. Ranked version (already sorted)
    ranked = combined_df.sort_values(['Track', 'RaceNumber', 'FinalScore'], 
                                     ascending=[True, True, False])
    color_code_predictions(ranked,
                          os.path.join(output_dir, "ranked_color.xlsx"))
    
    # 3. Top 3 picks per race
    top3 = ranked.groupby(['Track', 'RaceNumber']).head(3).reset_index(drop=True)
    color_code_predictions(top3,
                          os.path.join(output_dir, "top3_picks_color.xlsx"))
    
    print(f"\n✨ Created 3 color-coded Excel files in {output_dir}/")
    print("   • todays_form_color.xlsx - All dogs with color-coded rankings")
    print("   • ranked_color.xlsx - Sorted by race with color-coded rankings")
    print("   • top3_picks_color.xlsx - Only top 3 per race with color-coding")
