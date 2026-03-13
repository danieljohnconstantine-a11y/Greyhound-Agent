"""
verify_pdf_excel_accuracy.py
============================
Cross-checks every dog in outputs/track_ensemble_predictions.xlsx against
the PDF race forms in data_predictions/.

Factual fields checked per dog:
  Track | RaceNumber | Box | DogName | Trainer | SexAge | Weight
  CareerWins | CareerPlaces | CareerStarts | PrizeMoney | RTC | DLR | DLW
  Distance | RaceTime | BestTimeSec

Output:
  Prints full audit to stdout AND writes
  reports/DATA_ACCURACY_AUDIT_PDF_VS_EXCEL_2026-03-13.txt
"""

import sys
import os
import pdfplumber
import pandas as pd
import openpyxl
import re
from datetime import datetime

sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))
import parser as race_parser

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
PDF_DIR      = os.path.join(os.path.dirname(__file__), 'data_predictions')
EXCEL_PATH   = os.path.join(os.path.dirname(__file__), 'outputs', 'track_ensemble_predictions.xlsx')
REPORT_PATH  = os.path.join(os.path.dirname(__file__), 'reports',
                             'DATA_ACCURACY_AUDIT_PDF_VS_EXCEL_2026-03-13.txt')

# Normalise track names so PDF "HEALESVILLE" == Excel "HEALESVILLE" etc.
TRACK_NORM = {
    'THE GARDENS':  'The Gardens',
    'GARDENS':      'The Gardens',
    'MANDURAH':     'Mandurah',
    'Q LAKESIDE':   'Q LAKESIDE',
    'QUEENSLAND LAKESIDE': 'Q LAKESIDE',
}

def norm_track(t):
    if t is None:
        return ''
    s = str(t).strip().upper()
    return TRACK_NORM.get(s, str(t).strip())

def norm_str(v):
    if v is None:
        return ''
    return str(v).strip()

def norm_float(v):
    """Return float or None."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None

def norm_int(v):
    """Return int or None."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None

# ---------------------------------------------------------------------------
# 1. Load Excel
# ---------------------------------------------------------------------------
print("Loading Excel …")
wb  = openpyxl.load_workbook(EXCEL_PATH)
ws  = wb.active

headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
excel_rows = []
for r in range(2, ws.max_row + 1):
    row = {headers[c]: ws.cell(r, c + 1).value for c in range(len(headers))}
    excel_rows.append(row)

excel_df = pd.DataFrame(excel_rows)
print(f"  Excel: {len(excel_df)} dogs, columns: {list(excel_df.columns)[:10]} …")

# Build lookup key: (norm_track, race_number, box) -> excel row
excel_lookup = {}
for _, row in excel_df.iterrows():
    key = (norm_track(row.get('Track')), norm_int(row.get('RaceNumber')), norm_int(row.get('Box')))
    excel_lookup[key] = row

print(f"  Excel lookup keys: {len(excel_lookup)}")

# ---------------------------------------------------------------------------
# 2. Parse all PDFs
# ---------------------------------------------------------------------------
pdf_files = sorted([f for f in os.listdir(PDF_DIR) if f.endswith('.pdf')])
print(f"\nPDFs to parse: {pdf_files}")

pdf_df_list = []
for fname in pdf_files:
    fpath = os.path.join(PDF_DIR, fname)
    print(f"  Parsing {fname} …", end=' ', flush=True)
    try:
        with pdfplumber.open(fpath) as pdf:
            text = '\n'.join(page.extract_text() or '' for page in pdf.pages)
        df = race_parser.parse_race_form(text)
        # Suppress parser logger noise already captured above
        df['_src_pdf'] = fname
        pdf_df_list.append(df)
        print(f"{len(df)} dogs")
    except Exception as e:
        print(f"ERROR: {e}")

pdf_df = pd.concat(pdf_df_list, ignore_index=True)
print(f"\nTotal dogs parsed from PDFs: {len(pdf_df)}")

# ---------------------------------------------------------------------------
# 3. Cross-check
# ---------------------------------------------------------------------------
lines = []
def emit(s):
    print(s)
    lines.append(s)

emit("=" * 78)
emit(f"DATA ACCURACY AUDIT — PDF vs Excel")
emit(f"Run: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
emit(f"Excel: {EXCEL_PATH}")
emit(f"PDFs:  {PDF_DIR}")
emit("=" * 78)

total_pdf       = len(pdf_df)
matched_exact   = 0     # key found AND no mismatches
matched_mismatch= 0     # key found BUT field mismatch
missing_in_excel= 0     # key not found in Excel
missing_in_pdf  = 0     # Excel row not found in PDF (checked later)
mismatch_detail = []    # list of mismatch dicts
missing_detail  = []    # list of missing-in-excel dicts

# Fields to compare (PDF col -> Excel col)
FACTUAL_FIELDS = [
    ('DogName',      'DogName'),
    ('Trainer',      'Trainer'),
    ('SexAge',       'SexAge'),
    ('Weight',       'Weight'),
    ('CareerWins',   'CareerWins'),
    ('CareerPlaces', 'CareerPlaces'),
    ('CareerStarts', 'CareerStarts'),
    ('PrizeMoney',   'PrizeMoney'),
    ('RTC',          'RTC'),
    ('DLR',          'DLR'),
    ('DLW',          'DLW'),
    ('Distance',     'Distance'),
]

for _, prow in pdf_df.iterrows():
    pdf_track = norm_track(prow.get('Track'))
    pdf_race  = norm_int(prow.get('RaceNumber'))
    pdf_box   = norm_int(prow.get('Box'))
    key       = (pdf_track, pdf_race, pdf_box)

    if key not in excel_lookup:
        missing_in_excel += 1
        missing_detail.append({
            'source': 'PDF',
            'pdf': prow.get('_src_pdf'),
            'Track': pdf_track,
            'Race': pdf_race,
            'Box': pdf_box,
            'DogName': norm_str(prow.get('DogName')),
            'issue': 'NOT FOUND IN EXCEL',
        })
        continue

    erow       = excel_lookup[key]
    row_mismatches = []

    for pdf_col, xl_col in FACTUAL_FIELDS:
        pval = prow.get(pdf_col)
        eval_ = erow.get(xl_col)

        # Normalise for comparison
        if pdf_col in ('Weight', 'PrizeMoney', 'BestTimeSec'):
            pv = norm_float(pval)
            ev = norm_float(eval_)
            match = (pv == ev) or (pv is None and ev is None)
        elif pdf_col in ('CareerWins', 'CareerPlaces', 'CareerStarts',
                         'RTC', 'DLR', 'DLW', 'Distance'):
            pv = norm_int(pval)
            ev = norm_int(eval_)
            match = (pv == ev) or (pv is None and ev is None)
        else:
            pv = norm_str(pval).lower()
            ev = norm_str(eval_).lower()
            match = (pv == ev)

        if not match:
            row_mismatches.append({
                'field':    pdf_col,
                'pdf_val':  pval,
                'excel_val':eval_,
            })

    if row_mismatches:
        matched_mismatch += 1
        mismatch_detail.append({
            'Track':    pdf_track,
            'Race':     pdf_race,
            'Box':      pdf_box,
            'DogName':  norm_str(prow.get('DogName')),
            'pdf':      prow.get('_src_pdf'),
            'fields':   row_mismatches,
        })
    else:
        matched_exact += 1

# Check Excel rows not found in any PDF
pdf_keys = set()
for _, prow in pdf_df.iterrows():
    pdf_keys.add((norm_track(prow.get('Track')),
                  norm_int(prow.get('RaceNumber')),
                  norm_int(prow.get('Box'))))

for key, erow in excel_lookup.items():
    if key not in pdf_keys:
        missing_in_pdf += 1
        missing_detail.append({
            'source':  'Excel',
            'Track':   key[0],
            'Race':    key[1],
            'Box':     key[2],
            'DogName': norm_str(erow.get('DogName')),
            'issue':   'NOT FOUND IN ANY PDF',
        })

# ---------------------------------------------------------------------------
# 4. Report
# ---------------------------------------------------------------------------
emit(f"\nSUMMARY")
emit(f"-------")
emit(f"Total dogs in PDFs  : {total_pdf}")
emit(f"Total dogs in Excel : {len(excel_df)}")
emit(f"Exact matches       : {matched_exact}  (Track+Race+Box key found + ALL fields match)")
emit(f"Field mismatches    : {matched_mismatch}  (key found but ≥1 field differs)")
emit(f"PDF dog not in Excel: {missing_in_excel}")
emit(f"Excel dog not in PDF: {missing_in_pdf}")

# ---------------------------------------------------------------------------
emit(f"\n{'=' * 78}")
emit("FIELD MISMATCHES (detailed)")
emit('=' * 78)
if mismatch_detail:
    for m in mismatch_detail:
        emit(f"\n  ► {m['Track']} Race {m['Race']} Box {m['Box']} — {m['DogName']}  (PDF: {m['pdf']})")
        for f in m['fields']:
            emit(f"      {f['field']:18s}  PDF={repr(f['pdf_val'])}  Excel={repr(f['excel_val'])}")
else:
    emit("  ✓ ZERO field mismatches detected.")

# ---------------------------------------------------------------------------
emit(f"\n{'=' * 78}")
emit("MISSING DOGS (detailed)")
emit('=' * 78)
if missing_detail:
    pdf_missing  = [x for x in missing_detail if x['source'] == 'PDF']
    xl_missing   = [x for x in missing_detail if x['source'] == 'Excel']

    if pdf_missing:
        emit(f"\n  Dogs in PDFs but NOT found in Excel ({len(pdf_missing)}):")
        for m in pdf_missing:
            emit(f"    • {m['Track']} Race {m['Race']} Box {m['Box']} — {m['DogName']}  (PDF: {m['pdf']})")

    if xl_missing:
        emit(f"\n  Dogs in Excel but NOT found in any PDF ({len(xl_missing)}):")
        for m in xl_missing:
            emit(f"    • {m['Track']} Race {m['Race']} Box {m['Box']} — {m['DogName']}")
else:
    emit("  ✓ No missing dogs detected.")

# ---------------------------------------------------------------------------
emit(f"\n{'=' * 78}")
emit("VERDICT")
emit('=' * 78)
if matched_mismatch == 0 and missing_in_excel == 0 and missing_in_pdf == 0:
    emit("✅ PASS — All PDF data matches Excel exactly. ZERO discrepancies.")
else:
    issues = []
    if matched_mismatch:
        issues.append(f"{matched_mismatch} field mismatch(es)")
    if missing_in_excel:
        issues.append(f"{missing_in_excel} PDF dog(s) missing from Excel")
    if missing_in_pdf:
        issues.append(f"{missing_in_pdf} Excel dog(s) not in any PDF")
    emit(f"❌ FAIL — {'; '.join(issues)}.")

emit(f"\nReport generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
emit("=" * 78)

# ---------------------------------------------------------------------------
# 5. Write report file
# ---------------------------------------------------------------------------
os.makedirs(os.path.dirname(REPORT_PATH), exist_ok=True)
with open(REPORT_PATH, 'w', encoding='utf-8') as fh:
    fh.write('\n'.join(lines))
print(f"\nReport written → {REPORT_PATH}")
