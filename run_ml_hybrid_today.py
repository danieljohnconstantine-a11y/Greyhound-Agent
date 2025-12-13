"""
ML Hybrid Prediction System for Today's Races
Combines v4.4 rule-based scoring with Machine Learning predictions.
Generates separate Excel outputs for ML hybrid picks.

Usage:
    python run_ml_hybrid_today.py
    
Reads PDFs from: data_predictions/
Outputs to: outputs/ml_hybrid_picks.xlsx
"""

import sys
import os
import glob
import pandas as pd
import pdfplumber
from datetime import datetime

sys.path.insert(0, os.path.dirname(__file__))

# Fix encoding for Windows console to handle emoji characters
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from src.parser import parse_race_form
from src.features import compute_features
from src.scorer import score_race
from src.bet_worthy import detect_bet_worthy
from src.ml_predictor import GreyhoundMLPredictor
from src.excel_export import create_color_coded_outputs

def main():
    print("=" * 80)
    print("🤖 ML HYBRID PREDICTION SYSTEM")
    print("   v4.4 Rule-Based + Machine Learning")
    print("=" * 80)
    print()
    
    # Check for ML model
    model_path = 'models/greyhound_ml_v1.pkl'
    if not os.path.exists(model_path):
        print(f"❌ ML model not found at: {model_path}")
        print()
        print("   To train the ML model:")
        print("   1. Run train_ml.bat (Windows) or python train_ml_model.py")
        print("   2. Wait 2-3 minutes for training to complete")
        print("   3. Model will be saved to models/greyhound_ml_v1.pkl")
        print()
        print("   For now, use run_predictions_today.bat for v4.4-only predictions")
        return
    
    # Load ML model
    print("📥 Loading ML model...")
    try:
        predictor = GreyhoundMLPredictor(model_path)
        print("   ✅ ML model loaded successfully")
    except Exception as e:
        print(f"   ❌ Error loading ML model: {e}")
        return
    
    # Get PDFs from data_predictions folder
    pdf_folder = 'data_predictions'
    if not os.path.exists(pdf_folder):
        print(f"❌ Folder not found: {pdf_folder}")
        print(f"   Please create '{pdf_folder}' and add today's race PDFs")
        return
    
    pdf_files = glob.glob(os.path.join(pdf_folder, '*form.pdf'))
    if not pdf_files:
        print(f"❌ No PDF files found in {pdf_folder}/")
        print(f"   Please add today's race form PDFs to {pdf_folder}/")
        return
    
    print(f"📄 Found {len(pdf_files)} race PDF(s)")
    for pdf_file in pdf_files:
        print(f"   • {os.path.basename(pdf_file)}")
    print()
    
    # Process all PDFs with enhanced error tracking
    all_ml_picks = []
    all_v44_picks = []
    all_ml_predictions = []  # All ML predictions for every dog
    hybrid_bet_count = 0
    v44_bet_count = 0
    
    # Enhanced error tracking
    error_log = []
    total_pdfs = len(pdf_files)
    successful_pdfs = 0
    failed_pdfs = 0
    total_races = 0
    successful_races = 0
    failed_races = 0
    
    for pdf_idx, pdf_file in enumerate(pdf_files, 1):
        pdf_name = os.path.basename(pdf_file)
        print(f"🔍 Processing ({pdf_idx}/{total_pdfs}): {pdf_name}")
        
        try:
            # Extract text from PDF
            text = ""
            try:
                with pdfplumber.open(pdf_file) as pdf:
                    for page in pdf.pages:
                        text += page.extract_text() + "\n"
            except Exception as e:
                error_msg = f"PDF extraction failed for {pdf_name}: {type(e).__name__}: {str(e)}"
                error_log.append(error_msg)
                print(f"   ❌ {error_msg}")
                failed_pdfs += 1
                continue
            
            # Parse race form - returns a DataFrame with all dogs
            try:
                df_all = parse_race_form(text)
                if df_all is None or len(df_all) == 0:
                    error_msg = f"No races found in {pdf_name}"
                    error_log.append(error_msg)
                    print(f"   ⚠️  {error_msg}")
                    failed_pdfs += 1
                    continue
            except Exception as e:
                import traceback
                error_msg = f"Race parsing failed for {pdf_name}: {type(e).__name__}: {str(e)}"
                error_log.append(error_msg)
                print(f"   ❌ {error_msg}")
                print(f"   Full traceback:")
                traceback.print_exc()
                failed_pdfs += 1
                continue
            
            # Convert DLR to numeric to avoid type errors
            df_all["DLR"] = pd.to_numeric(df_all["DLR"], errors="coerce")
            
            # Apply enhanced scoring to all dogs
            try:
                df_all = compute_features(df_all)
            except Exception as e:
                error_msg = f"{pdf_name}: Feature computation failed - {type(e).__name__}: {str(e)}"
                error_log.append(error_msg)
                print(f"   ❌ {error_msg}")
                failed_pdfs += 1
                continue
            
            print(f"   ✅ Parsed {len(df_all)} dogs")
            successful_pdfs += 1
            
            # Process each race - group by Track and RaceNumber
            races = df_all.groupby(['Track', 'RaceNumber'])
            for (track, race_num), df_race in races:
                total_races += 1
                
                try:
                    # Score with v4.4 system
                    try:
                        df_scored = score_race(df_race, track)
                        rule_based_scores = df_scored['FinalScore']
                    except Exception as e:
                        error_msg = f"{pdf_name} Race {race_num}: v4.4 scoring failed - {type(e).__name__}: {str(e)}"
                        error_log.append(error_msg)
                        print(f"      ❌ {error_msg}")
                        failed_races += 1
                        continue
                    
                    # Detect TIER0 with v4.4
                    try:
                        bet_worthy_info = detect_bet_worthy(df_scored, track=track)
                        v4_4_tier = bet_worthy_info.get('tier', 'NONE')
                        v4_4_recommended = bet_worthy_info.get('recommended_box')
                    except Exception as e:
                        import traceback
                        error_msg = f"{pdf_name} Race {race_num}: Bet-worthy detection failed - {type(e).__name__}: {str(e)}"
                        error_log.append(error_msg)
                        print(f"      ❌ {error_msg}")
                        print(f"         DataFrame types: {df_scored.dtypes.to_dict() if hasattr(df_scored, 'dtypes') else 'N/A'}")
                        print(f"         Full traceback:")
                        traceback.print_exc()
                        failed_races += 1
                        continue
                    
                    # Get ML hybrid prediction
                    try:
                        hybrid_result = predictor.hybrid_predict(
                            df_race,
                            rule_based_scores,
                            tier0_threshold=18,  # v4.4 TIER0 margin threshold
                            ml_threshold=75      # ML confidence threshold (75%)
                        )
                        
                        # Store ALL ML predictions (for every dog in race)
                        all_preds = hybrid_result.get('all_predictions')
                        if all_preds is not None and not all_preds.empty:
                            for _, pred_row in all_preds.iterrows():
                                ml_pred = {
                                    'Track': track,
                                    'RaceNumber': race_num,
                                    'Box': pred_row['Box'],
                                    'DogName': pred_row.get('DogName', 'Unknown'),
                                    'ML_Confidence': pred_row['ML_Confidence'],
                                    'v4.4_Score': pred_row['RuleBased_Score'],
                                    'Rank': None  # Will be set later after sorting
                                }
                                all_ml_predictions.append(ml_pred)
                        
                        # Store ML hybrid pick if it qualifies
                        if hybrid_result['tier'] == 'HYBRID_TIER0':
                            dog = df_race[df_race['Box'] == hybrid_result['recommended_box']].iloc[0]
                            ml_pick = {
                                'Track': track,
                                'RaceNumber': race_num,
                                'Box': hybrid_result['recommended_box'],
                                'DogName': dog.get('DogName', 'Unknown'),
                                'v4.4_Score': hybrid_result['rule_based_score'],
                                'v4.4_Margin': hybrid_result['margin_pct'],
                                'ML_Confidence': hybrid_result['ml_confidence'],
                                'Tier': 'HYBRID_TIER0',
                                'Recommendation': '✅ STRONG BET (ML + v4.4 Agree)',
                                'ExpectedWinRate': '35-40%'
                            }
                            all_ml_picks.append(ml_pick)
                            hybrid_bet_count += 1
                            print(f"      ✅ Race {race_num}: HYBRID BET - Box {hybrid_result['recommended_box']} ({dog.get('DogName', 'Unknown')})")
                        else:
                            print(f"      ⚠️  Race {race_num}: No hybrid bet (ML: {hybrid_result['ml_confidence']:.1f}%, v4.4: {hybrid_result['margin_pct']:.1f}%)")
                        
                        # Also store v4.4-only pick for comparison
                        if v4_4_recommended is not None:
                            dog_v44 = df_scored[df_scored['Box'] == v4_4_recommended].iloc[0]
                            v44_pick = {
                                'Track': track,
                                'RaceNumber': race_num,
                                'Box': v4_4_recommended,
                                'DogName': dog_v44.get('DogName', 'Unknown'),
                                'FinalScore': rule_based_scores.max(),
                                'Tier': v4_4_tier,
                                'Margin': bet_worthy_info.get('margin_pct', 0),
                                'Recommendation': 'v4.4 Pick',
                                'ExpectedWinRate': '28-30%'
                            }
                            all_v44_picks.append(v44_pick)
                            v44_bet_count += 1
                        
                        successful_races += 1
                        
                    except Exception as e:
                        error_msg = f"{pdf_name} Race {race_num}: ML prediction failed - {type(e).__name__}: {str(e)}"
                        error_log.append(error_msg)
                        print(f"      ❌ {error_msg}")
                        failed_races += 1
                        continue
                        
                except Exception as e:
                    error_msg = f"{pdf_name} Race {race_num}: Unexpected error - {type(e).__name__}: {str(e)}"
                    error_log.append(error_msg)
                    print(f"      ❌ {error_msg}")
                    failed_races += 1
                    continue
                
        except Exception as e:
            error_msg = f"{pdf_name}: Unexpected PDF-level error - {type(e).__name__}: {str(e)}"
            error_log.append(error_msg)
            print(f"   ❌ {error_msg}")
            failed_pdfs += 1
            continue
    
    # Print error summary
    print()
    print("=" * 80)
    print("📊 PROCESSING SUMMARY")
    print("=" * 80)
    print(f"   PDFs: {successful_pdfs}/{total_pdfs} successful, {failed_pdfs} failed")
    print(f"   Races: {successful_races}/{total_races} successful, {failed_races} failed")
    
    if error_log:
        print()
        print("=" * 80)
        print("⚠️  ERROR LOG - All Issues Encountered:")
        print("=" * 80)
        for idx, error in enumerate(error_log, 1):
            print(f"{idx}. {error}")
        print()
        print(f"Total errors: {len(error_log)}")
    else:
        print(f"   ✅ No errors encountered!")
    print()
    
    print()
    print("=" * 80)
    print("📊 PREDICTION SUMMARY")
    print("=" * 80)
    print(f"   Total races analyzed: {len(all_v44_picks)}")
    print(f"   v4.4 picks (TIER0): {v44_bet_count}")
    print(f"   ML Hybrid picks (Both systems agree): {hybrid_bet_count}")
    print(f"   Selectivity: {(hybrid_bet_count/max(len(all_v44_picks), 1))*100:.1f}% of races")
    print()
    
    # Create outputs folder
    os.makedirs('outputs', exist_ok=True)
    
    # Save ML hybrid picks to Excel
    if all_ml_picks:
        df_ml_picks = pd.DataFrame(all_ml_picks)
        excel_path = 'outputs/ml_hybrid_picks.xlsx'
        
        # Create Excel with formatting
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_ml_picks.to_excel(writer, sheet_name='ML Hybrid Picks', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['ML Hybrid Picks']
            
            # Auto-adjust column widths
            for idx, col in enumerate(df_ml_picks.columns, 1):
                max_length = max(
                    df_ml_picks[col].astype(str).apply(len).max(),
                    len(col)
                )
                worksheet.column_dimensions[chr(64 + idx)].width = min(max_length + 2, 50)
            
            # Add header formatting
            from openpyxl.styles import Font, PatternFill, Alignment
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        print(f"✅ ML Hybrid picks saved to: {excel_path}")
        print()
        print("🎯 ML HYBRID PICKS (Both systems strongly agree):")
        print("-" * 80)
        for pick in all_ml_picks:
            print(f"   🤖 {pick['Track']} Race {pick['RaceNumber']}: Box {pick['Box']} - {pick['DogName']}")
            print(f"      v4.4 Score: {pick['v4.4_Score']:.1f} (Margin: {pick['v4.4_Margin']:.1f}%)")
            print(f"      ML Confidence: {pick['ML_Confidence']:.1f}%")
            print(f"      Expected Win Rate: {pick['ExpectedWinRate']}")
            print()
    else:
        print("ℹ️  No ML hybrid picks today")
        print("   (Waiting for both v4.4 AND ML to strongly agree)")
        print()
    
    # Save ALL ML predictions to Excel (sorted by confidence)
    if all_ml_predictions:
        df_all_ml = pd.DataFrame(all_ml_predictions)
        # Sort by ML confidence (highest to lowest)
        df_all_ml = df_all_ml.sort_values('ML_Confidence', ascending=False).reset_index(drop=True)
        
        # Add rank column (1 = highest confidence)
        df_all_ml['Rank'] = range(1, len(df_all_ml) + 1)
        
        # Reorder columns
        df_all_ml = df_all_ml[['Rank', 'Track', 'RaceNumber', 'Box', 'DogName', 'ML_Confidence', 'v4.4_Score']]
        
        excel_all_path = 'outputs/ml_all_predictions.xlsx'
        
        # Create Excel with formatting
        with pd.ExcelWriter(excel_all_path, engine='openpyxl') as writer:
            df_all_ml.to_excel(writer, sheet_name='All ML Predictions', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['All ML Predictions']
            
            # Auto-adjust column widths
            for idx, col in enumerate(df_all_ml.columns, 1):
                max_length = max(
                    df_all_ml[col].astype(str).apply(len).max(),
                    len(col)
                )
                worksheet.column_dimensions[chr(64 + idx)].width = min(max_length + 2, 50)
            
            # Add header formatting
            from openpyxl.styles import Font, PatternFill, Alignment
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Highlight top predictions (ML confidence >= 75%)
            for row in range(2, len(df_all_ml) + 2):
                confidence_cell = worksheet.cell(row=row, column=6)  # ML_Confidence column
                if float(confidence_cell.value) >= 75:
                    for col in range(1, 8):
                        cell = worksheet.cell(row=row, column=col)
                        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        
        print(f"✅ All ML predictions saved to: {excel_all_path}")
        print(f"   Total predictions: {len(df_all_ml)} (sorted by ML confidence)")
        print(f"   Top ML pick overall: {df_all_ml.iloc[0]['Track']} R{df_all_ml.iloc[0]['RaceNumber']} "
              f"Box {df_all_ml.iloc[0]['Box']} ({df_all_ml.iloc[0]['ML_Confidence']:.1f}%)")
        print()
    
    # Save v4.4 picks for comparison
    if all_v44_picks:
        df_v44 = pd.DataFrame(all_v44_picks)
        df_v44.to_csv('outputs/v44_picks_comparison.csv', index=False)
        print(f"📊 v4.4 picks saved for comparison: outputs/v44_picks_comparison.csv")
        print()
    
    print("=" * 80)
    print("✅ ML HYBRID ANALYSIS COMPLETE")
    print("=" * 80)
    print()
    print("📁 OUTPUT FILES:")
    print("   1. outputs/ml_hybrid_picks.xlsx - High-confidence bets (ML + v4.4 agree)")
    print("   2. outputs/ml_all_predictions.xlsx - ALL ML predictions ranked by confidence")
    print("   3. outputs/v44_picks_comparison.csv - All v4.4 picks for comparison")
    print()
    print("💡 KEY INSIGHTS:")
    print("   • ML Hybrid is MORE SELECTIVE but MORE ACCURATE")
    print("   • Only bets when v4.4 (18%+ margin) AND ML (75%+ confidence) agree")
    print("   • Target: 35-40% win rate vs 28-30% v4.4 alone")
    print("   • Fewer bets = higher quality picks")
    print("   • ml_all_predictions.xlsx shows EVERY dog ranked by ML confidence")
    print()
    print(f"📁 Check outputs folder for:")
    if all_ml_picks:
        print(f"   • ml_hybrid_picks.xlsx - Your ML hybrid bets (Excel with formatting)")
    print(f"   • v44_picks_comparison.csv - All v4.4 picks for reference")
    print()

if __name__ == "__main__":
    main()
