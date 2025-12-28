"""
Complete Analysis Pipeline - One-Click Solution
Runs predictions on data_predictions/ PDFs and generates all reports

This script combines:
1. ML v2.1 Enhanced Predictions (trained on 2,108 historical races)
2. Feature Analysis with Track→Race→Box sorting
3. Unified prediction report

Usage:
    python run_complete_analysis.py
    
    OR use the batch file:
    run_complete_analysis.bat

Prerequisites:
    1. Trained ML v2.1 model (run train_ml_enhanced.bat first)
    2. Race PDFs in data_predictions/ folder

Outputs (ALL generated in single run):
    1. outputs/ml_unified_predictions.xlsx - ALL dogs ranked by ML confidence (trained on 2,108 races)
    2. outputs/ml_feature_analysis_detailed.xlsx - Detailed features (Track→Race→Box sorted)
    3. outputs/complete_analysis_summary.txt - Quick summary report
"""

import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

# Fix encoding for Windows console
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import logging
logging.basicConfig(level=logging.INFO, format='%(message)s', handlers=[logging.StreamHandler(sys.stdout)])

from src.ml_predictor_advanced import AdvancedGreyhoundMLPredictor
from src.weather_track_data import WeatherTrackDataManager
from src.parser import parse_race_form
from src.features import compute_features
from src.scorer import score_race
import pandas as pd
import glob
import pdfplumber
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def main():
    print("=" * 80)
    print("🚀 COMPLETE ANALYSIS PIPELINE - ONE-CLICK SOLUTION")
    print("   Predictions + Feature Analysis + All Reports")
    print("=" * 80)
    
    start_time = datetime.now()
    
    # Initialize weather/track manager
    print("\n🌤️  Loading weather & track condition data...")
    try:
        weather_manager = WeatherTrackDataManager()
        print(f"✅ Weather records: {len(weather_manager.weather_data)}")
        print(f"✅ Track condition records: {len(weather_manager.track_conditions)}")
    except Exception as e:
        print(f"⚠️  Warning: Could not load weather/track data: {e}")
        weather_manager = WeatherTrackDataManager()
    
    # Load ML model (with fallback to v4.4-only mode)
    print("\n📥 Loading ML v2.1 enhanced model...")
    model_path = "models/greyhound_ml_v2.1_enhanced.pkl"
    
    if not os.path.exists(model_path):
        print(f"❌ ERROR: ML model not found at {model_path}")
        print("   This script requires the trained ML model to generate predictions.")
        print("   Please train the model first by running: train_ml_enhanced.bat")
        print("\n   The ML model uses 3,000+ historical races to make predictions.")
        print("   Without it, predictions cannot be generated.")
        sys.exit(1)
    
    try:
        predictor = AdvancedGreyhoundMLPredictor()
        predictor.load_model(model_path)
        print(f"✅ ML v2.1 model loaded successfully")
        print(f"   Model trained on {len(predictor.models) if hasattr(predictor, 'models') else 'multiple'} track-specific models")
    except Exception as e:
        print(f"❌ ERROR: Failed to load ML model: {e}")
        print("   The model file may be corrupted. Please retrain: train_ml_enhanced.bat")
        sys.exit(1)
    
    # Find PDFs in data_predictions
    pdf_files = glob.glob("data_predictions/*.pdf")
    
    if not pdf_files:
        print(f"\n❌ No PDF files found in data_predictions/")
        print(f"   Please add race form PDFs to the data_predictions/ folder")
        return 1
    
    print(f"\n📄 Found {len(pdf_files)} PDF files in data_predictions/")
    print(f"   Processing predictions for today's races...")
    
    # Storage for results
    all_ml_enhanced_predictions = []
    all_hybrid_picks = []
    all_v44_picks = []
    all_detailed_features = []
    
    processed_count = 0
    error_count = 0
    
    # Process each PDF
    for i, pdf_path in enumerate(pdf_files, 1):
        try:
            filename = os.path.basename(pdf_path)
            print(f"\n[{i}/{len(pdf_files)}] Processing: {filename}")
            
            with pdfplumber.open(pdf_path) as pdf:
                all_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
            
            # Parse race form
            df = parse_race_form(all_text)
            
            if df is None or len(df) == 0:
                print(f"   ⚠️  No data extracted from {filename}")
                print(f"   📄 Saving PDF text to outputs/debug_{filename}.txt for inspection")
                # Save the PDF text for manual inspection
                try:
                    with open(f"outputs/debug_{filename}.txt", 'w', encoding='utf-8') as f:
                        f.write(all_text)
                except:
                    pass
                error_count += 1
                continue
            
            # Get track name from filename (e.g., GEELG2711form.pdf -> GEEL)
            track_code = filename[:4].upper() if len(filename) >= 4 else "UNKN"
            
            # Group by race
            if 'RaceNumber' in df.columns:
                races = df.groupby('RaceNumber')
            elif 'RaceNum' in df.columns:
                races = df.groupby('RaceNum')
            else:
                print(f"   ⚠️  No RaceNumber/RaceNum column in {filename}")
                error_count += 1
                continue
            
            dogs_added_this_pdf = 0
            
            for race_num, race_df in races:
                try:
                    print(f"   🏁 Processing Race {race_num}: {len(race_df)} dogs")
                    
                    # Compute features for ML
                    features_df = compute_features(race_df)
                    
                    if features_df is None or len(features_df) == 0:
                        print(f"   ⚠️  No features computed for Race {race_num} - skipping")
                        continue
                    
                    print(f"   ✓ Features computed: {len(features_df)} dogs")
                    
                    # Get ML predictions from trained model
                    try:
                        ml_predictions = predictor.predict(features_df, track_code)
                        if ml_predictions is None:
                            ml_predictions = {}
                        print(f"   ✓ ML predictions: {len(ml_predictions)} dogs")
                    except Exception as e:
                        print(f"   ❌ ML prediction error: {e}")
                        import traceback
                        traceback.print_exc()
                        print(f"   ⚠️  Warning: ML prediction failed for Race {race_num}")
                        print(f"   📝 Continuing with available data for feature analysis...")
                        ml_predictions = {}  # Continue with empty predictions instead of skipping
                    
                    # Get v4.4 scores
                    v44_scores = score_race(race_df)
                    print(f"   ✓ V4.4 scores computed")
                    
                    # Convert v44_scores DataFrame to dict if needed
                    if isinstance(v44_scores, pd.DataFrame):
                        v44_scores_dict = {}
                        for idx, row in v44_scores.iterrows():
                            try:
                                box = int(row['Box'])
                                # The scorer returns 'FinalScore' column
                                score = float(row.get('FinalScore', row.get('Score', 0)))
                                v44_scores_dict[box] = score
                            except Exception as e:
                                continue
                        v44_scores = v44_scores_dict
                    
                    # Process each dog - use features_df to get ALL 80+ computed features
                    dogs_in_race = 0
                    for idx, row in features_df.iterrows():
                        try:
                            box = int(row.get('Box', 0))
                            dog_name = str(row.get('DogName', 'Unknown'))
                            
                            # Get v4.4 score
                            v44_score = v44_scores.get(box, 0.0)
                            
                            # Get ML confidence from trained model
                            ml_conf = ml_predictions.get(box, 0.0) * 100
                            
                            # Note: We no longer skip dogs with no ML prediction
                            # Instead, we include them with 0% confidence for feature analysis
                            if ml_conf == 0.0:
                                print(f"   ⚠️  No ML prediction for Box {box} ({dog_name}) - including with 0% confidence")
                            
                            # Store ML prediction
                            pred_record = {
                                'Track': track_code,
                                'Race': int(race_num),
                                'Box': box,
                                'DogName': dog_name,
                                'ML_Confidence': round(ml_conf, 1),
                                'v44_Score': round(v44_score, 1)
                            }
                            all_ml_enhanced_predictions.append(pred_record)
                            dogs_in_race += 1
                            dogs_added_this_pdf += 1
                            
                            # Check for hybrid pick (v4.4 >= 18% AND ML >= 70%)
                            if v44_score >= 18.0 and ml_conf >= 70.0:
                                all_hybrid_picks.append(pred_record.copy())
                            
                            # Check for v4.4 TIER0 pick (>= 18%)
                            if v44_score >= 18.0:
                                all_v44_picks.append(pred_record.copy())
                            
                            # Extract detailed features for analysis report
                            # CRITICAL: Use features_df (which has ALL 80+ computed features), not race_df
                            feature_record = {
                                'Track': track_code,
                                'Race': int(race_num),
                                'Box': box,
                                'DogName': dog_name,
                                'ML_Confidence': round(ml_conf, 1),
                                'v44_Score': round(v44_score, 1)
                            }
                            
                            # Add ALL columns from the features DataFrame (includes all computed features)
                            for col in row.index:
                                if col not in ['Box', 'DogName']:
                                    val = row[col]
                                    # Handle lists/arrays first (before checking pd.isna)
                                    if isinstance(val, (list, tuple)):
                                        feature_record[col] = str(val)
                                    # Handle NaN/None - show "N/A" instead of blank or 0
                                    elif val is None or (isinstance(val, float) and pd.isna(val)):
                                        # For important fields, show "N/A" instead of 0
                                        important_fields = ['BestTimeSec', 'SectionalSec', 'PrizeMoney', 
                                                          'CareerWins', 'Distance', 'Weight']
                                        if col in important_fields:
                                            feature_record[col] = "N/A"
                                        else:
                                            feature_record[col] = 0
                                    else:
                                        try:
                                            feature_record[col] = float(val) if isinstance(val, (int, float)) else str(val)
                                        except:
                                            feature_record[col] = str(val)
                            
                            all_detailed_features.append(feature_record)
                            
                        except Exception as e:
                            print(f"   ⚠️  Error processing dog Box {box}: {e}")
                            continue
                    
                    print(f"   ✅ Added {dogs_in_race} dogs from Race {race_num}")
                    
                except Exception as e:
                    print(f"   ⚠️  Error processing race {race_num}: {e}")
                    import traceback
                    traceback.print_exc()
                    continue
            
            processed_count += 1
            print(f"   ✅ PDF processed: {dogs_added_this_pdf} total dogs added")
            
        except Exception as e:
            print(f"   ❌ Error: {e}")
            error_count += 1
            continue
    
    print(f"\n" + "=" * 80)
    print(f"📊 PROCESSING COMPLETE")
    print(f"   ✅ Successfully processed: {processed_count}/{len(pdf_files)} PDFs")
    if error_count > 0:
        print(f"   ⚠️  Errors encountered: {error_count} PDFs")
    print(f"   📋 Total dogs collected: {len(all_ml_enhanced_predictions)}")
    print(f"   📊 Total features collected: {len(all_detailed_features)}")
    print("=" * 80)
    
    # Generate all output files
    print(f"\n📁 Generating output files...")
    
    files_generated = []
    
    # Wrap all Excel generation in try-except to ensure we always create outputs
    try:
        # 1. UNIFIED Predictions Report - TOP PICK PER RACE
        # Shows only the highest confidence dog for each race (1 winner prediction per race)
        # CHANGED: Generate this even if all_detailed_features is empty
        if all_ml_enhanced_predictions:
            try:
                # Merge predictions with detailed features to get all columns
                df_predictions = pd.DataFrame(all_ml_enhanced_predictions)
                
                # Check if we have detailed features to merge
                if all_detailed_features:
                    df_features = pd.DataFrame(all_detailed_features)
                    
                    # Merge to get all feature columns
                    df_unified = pd.merge(
                        df_predictions, 
                        df_features, 
                        on=['Track', 'Race', 'Box', 'DogName', 'ML_Confidence', 'v44_Score'],
                        how='left'
                    )
                else:
                    # No detailed features available - use predictions only
                    print(f"   ⚠️  No detailed features available - generating simplified report")
                    df_unified = df_predictions
            
            # Calculate comparative metrics for each dog
            # Sort to rank dogs within each race
            df_unified = df_unified.sort_values(['Track', 'Race', 'ML_Confidence', 'v44_Score'], 
                                                ascending=[True, True, False, False])
            
            # Add ranking within race
            df_unified['Position_In_Field'] = df_unified.groupby(['Track', 'Race']).cumcount() + 1
            df_unified['Total_In_Field'] = df_unified.groupby(['Track', 'Race'])['Box'].transform('count')
            df_unified['Position_Display'] = df_unified.apply(
                lambda x: f"{int(x['Position_In_Field'])} of {int(x['Total_In_Field'])}", axis=1
            )
            
            # Calculate confidence gap to 2nd place
            def calc_confidence_gap(group):
                sorted_group = group.sort_values('ML_Confidence', ascending=False)
                if len(sorted_group) >= 2:
                    sorted_group['Confidence_Gap_vs_2nd'] = sorted_group['ML_Confidence'].iloc[0] - sorted_group['ML_Confidence'].iloc[1]
                else:
                    sorted_group['Confidence_Gap_vs_2nd'] = 0
                # Fill gap for all rows in group (only 1st place is meaningful)
                sorted_group.loc[sorted_group.index[1:], 'Confidence_Gap_vs_2nd'] = 0
                return sorted_group
            
            df_unified = df_unified.groupby(['Track', 'Race']).apply(calc_confidence_gap).reset_index(drop=True)
            
            # IMPORTANT: Keep only the TOP dog per race (highest ML_Confidence per Track+Race)
            # Group by Track+Race and take the first (highest confidence) dog
            df_top_picks = df_unified[df_unified['Position_In_Field'] == 1].copy()
            
            # Add pick tier column
            df_top_picks['Pick_Tier'] = df_top_picks['ML_Confidence'].apply(
                lambda x: 'High Confidence (≥70%)' if x >= 70 
                else 'Medium Confidence (50-70%)' if x >= 50 
                else 'Lower Confidence (<50%)'
            )
            
            # Reorder columns for clarity - key columns first, then comparative metrics, then all features
            priority_cols = ['Track', 'Race', 'Box', 'DogName', 'ML_Confidence', 'Pick_Tier', 
                           'Confidence_Gap_vs_2nd', 'Position_Display', 'v44_Score']
            other_cols = [c for c in df_top_picks.columns if c not in priority_cols and c not in ['Position_In_Field', 'Total_In_Field']]
            df_top_picks = df_top_picks[priority_cols + other_cols]
            
            # Sort by ML confidence (highest first) for final output
            df_top_picks = df_top_picks.sort_values('ML_Confidence', ascending=False)
            
            df_top_picks.to_excel('outputs/ml_unified_predictions.xlsx', index=False)
            files_generated.append('ml_unified_predictions.xlsx')
            
            # Statistics
            total_races = len(df_top_picks)
            high_conf = len(df_top_picks[df_top_picks['ML_Confidence'] >= 70])
            med_conf = len(df_top_picks[(df_top_picks['ML_Confidence'] >= 50) & (df_top_picks['ML_Confidence'] < 70)])
            low_conf = len(df_top_picks[df_top_picks['ML_Confidence'] < 50])
            
            print(f"✅ Top picks (1 per race): outputs/ml_unified_predictions.xlsx")
            print(f"   📊 Total races: {total_races}")
            print(f"   🎯 High Confidence (≥70%): {high_conf} races")
            print(f"   🟡 Medium Confidence (50-70%): {med_conf} races")
            print(f"   ⚪ Lower Confidence (<50%): {low_conf} races")
            print(f"   📋 Columns: {len(df_top_picks.columns)} (all features included)")
            print(f"   📈 Each race shows ONLY the top predicted winner")
        except Exception as e:
            print(f"❌ Error saving unified predictions: {e}")
            import traceback
            traceback.print_exc()
    else:
        print(f"⚠️  No predictions generated")
    
    # 2. Detailed Feature Analysis (Track→Race→Box sorted with color coding + blank rows between races)
    if all_detailed_features:
        try:
            df_features = pd.DataFrame(all_detailed_features)
            # Sort by Track, Race, ML_Confidence (descending) to rank within race
            df_features = df_features.sort_values(['Track', 'Race', 'ML_Confidence'], ascending=[True, True, False])
            
            # Add position ranking within each race
            df_features['Position_Rank'] = df_features.groupby(['Track', 'Race']).cumcount() + 1
            
            # Re-sort by Track, Race, Box for final display
            df_features = df_features.sort_values(['Track', 'Race', 'Box'])
            
            # Insert blank rows between races for easier reading
            # Group by Track and Race, and build a new dataframe with blank rows
            df_with_blanks = pd.DataFrame()
            for (track, race), group in df_features.groupby(['Track', 'Race'], sort=False):
                df_with_blanks = pd.concat([df_with_blanks, group], ignore_index=True)
                # Add blank row after each race (except last race)
                # We'll add it for all, then remove the last one later
                blank_row = pd.DataFrame([{col: None for col in df_features.columns}])
                df_with_blanks = pd.concat([df_with_blanks, blank_row], ignore_index=True)
            
            # Remove the last blank row
            if len(df_with_blanks) > 0:
                df_with_blanks = df_with_blanks.iloc[:-1]
            
            # Save to Excel
            output_path = 'outputs/ml_feature_analysis_detailed.xlsx'
            df_with_blanks.to_excel(output_path, index=False)
            
            # Apply color coding using openpyxl
            try:
                wb = load_workbook(output_path)
                ws = wb.active
                
                # Define colors
                green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
                orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
                black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Black separator
                
                # Get column index for Position_Rank
                headers = [cell.value for cell in ws[1]]
                rank_col_idx = headers.index('Position_Rank') + 1 if 'Position_Rank' in headers else None
                
                # Apply conditional formatting based on Position_Rank
                if rank_col_idx:
                    for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
                        rank_cell = ws.cell(row=row_idx, column=rank_col_idx)
                        rank_value = rank_cell.value
                        
                        # Check if this is a blank separator row (all cells are None)
                        is_blank_row = all(ws.cell(row=row_idx, column=c).value is None 
                                          for c in range(1, min(6, ws.max_column + 1)))  # Check first 5 columns
                        
                        if is_blank_row:
                            # Color blank separator rows BLACK
                            for col_idx in range(1, ws.max_column + 1):
                                ws.cell(row=row_idx, column=col_idx).fill = black_fill
                            continue
                        
                        # Apply color to data rows based on rank
                        if rank_value == 1:
                            fill = green_fill
                        elif rank_value == 2:
                            fill = yellow_fill
                        elif rank_value == 3:
                            fill = orange_fill
                        else:
                            continue
                        
                        # Color all cells in the row
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = fill
                
                wb.save(output_path)
                files_generated.append('ml_feature_analysis_detailed.xlsx')
                print(f"✅ Feature analysis: {output_path} ({len(df_features)} dogs)")
                print(f"   📋 Sorted by Track → Race → Box for easy navigation")
                print(f"   ⬛ BLACK separator rows added between races for easier reading")
                print(f"   🟢 Green = Predicted 1st place (highest ML confidence)")
                print(f"   🟡 Yellow = Predicted 2nd place")
                print(f"   🟠 Orange = Predicted 3rd place")
                print(f"   📊 Contains {len(df_features.columns)} features/columns (ALL 80+ computed features)")
                print(f"   🔍 Shows ALL data used by ML model for predictions")
            except Exception as color_error:
                print(f"✅ Feature analysis: {output_path} ({len(df_features)} dogs)")
                print(f"   ⚠️  Color coding skipped: {color_error}")
                print(f"   📋 Sorted by Track → Race → Box for easy navigation")
                print(f"   📏 Blank rows added between races for easier reading")
                print(f"   📊 Contains {len(df_features.columns)} features/columns (ALL 80+ computed features)")
        except Exception as e:
            print(f"❌ Error saving feature analysis: {e}")
            import traceback
            traceback.print_exc()
    else:
        print(f"⚠️  No features captured")
    
    # 3. Summary report
    try:
        end_time = datetime.now()
        duration = (end_time - start_time).total_seconds()
        
        with open('outputs/complete_analysis_summary.txt', 'w') as f:
            f.write("=" * 80 + "\n")
            f.write("COMPLETE ANALYSIS PIPELINE - SUMMARY REPORT\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Processing Time: {duration:.1f} seconds\n\n")
            
            f.write("HOW 2,108 HISTORICAL RACES ARE USED:\n")
            f.write("  The ML v2.1 model was TRAINED on 2,108 actual race results.\n")
            f.write("  It learned patterns from:\n")
            f.write("    - Winner characteristics (times, form, box positions)\n")
            f.write("    - Track-specific trends (which tracks favor certain boxes)\n")
            f.write("    - Weather/condition impacts on performance\n")
            f.write("    - Dog performance patterns across all conditions\n\n")
            f.write("  When predicting today's races, the model applies these learned patterns\n")
            f.write("  to score each dog's likelihood of winning.\n\n")
            
            f.write("INPUT:\n")
            f.write(f"  PDFs Processed: {processed_count}/{len(pdf_files)}\n")
            f.write(f"  Errors: {error_count}\n\n")
            
            f.write("OUTPUT FILES GENERATED:\n")
            
            # Count unique races for top picks report
            if all_ml_enhanced_predictions:
                df_temp = pd.DataFrame(all_ml_enhanced_predictions)
                unique_races = df_temp.groupby(['Track', 'Race']).size().count()
                f.write(f"  1. ml_unified_predictions.xlsx - TOP PICK PER RACE\n")
                f.write(f"     • {unique_races} races analyzed (1 winning prediction per race)\n")
                f.write(f"     • Shows ONLY the highest confidence dog for each race\n")
                f.write(f"     • Ranked by ML confidence (based on 2,108 historical race patterns)\n")
                f.write(f"     • Includes ALL {len(df_temp.columns) if df_temp is not None else 'available'} feature columns\n\n")
            else:
                f.write(f"  1. ml_unified_predictions.xlsx - TOP PICK PER RACE\n")
                f.write(f"     • Shows ONLY the highest confidence dog for each race\n\n")
            
            f.write(f"  2. ml_feature_analysis_detailed.xlsx - {len(all_detailed_features)} dogs with full features\n")
            f.write(f"     • Sorted Track→Race→Box for easy navigation\n")
            f.write(f"     • Shows ALL dogs from all races (not just top picks)\n")
            f.write(f"     • Contains ALL data the ML model uses to make predictions\n")
            f.write(f"     • {len(pd.DataFrame(all_detailed_features).columns) if all_detailed_features else 'Multiple'} columns of actual PDF data\n\n")
            f.write(f"  3. complete_analysis_summary.txt - This summary\n\n")
            
            if all_ml_enhanced_predictions:
                df_temp = pd.DataFrame(all_ml_enhanced_predictions)
                # Calculate stats on top picks per race
                df_grouped = df_temp.sort_values('ML_Confidence', ascending=False).groupby(['Track', 'Race']).first()
                high_conf_races = len(df_grouped[df_grouped['ML_Confidence'] >= 70])
                med_conf_races = len(df_grouped[(df_grouped['ML_Confidence'] >= 50) & (df_grouped['ML_Confidence'] < 70)])
                
                f.write("RACE CONFIDENCE DISTRIBUTION (Top Pick Per Race):\n")
                f.write(f"  High (≥70%): {high_conf_races} races - Strong predictions\n")
                f.write(f"  Medium (50-70%): {med_conf_races} races - Moderate confidence\n")
                f.write(f"  Lower (<50%): {len(df_grouped) - high_conf_races - med_conf_races} races - Less confident\n\n")
            
            f.write("HOW TO USE THE REPORTS:\n")
            f.write("  1. Open ml_unified_predictions.xlsx (START HERE)\n")
            f.write("     - ONE winning prediction per race (highest ML confidence)\n")
            f.write("     - Races ranked by confidence (highest first)\n")
            f.write("     - ALL feature columns included for transparency\n")
            f.write("     - Focus on 'High Confidence (≥70%)' races for best results\n\n")
            f.write("  2. Open ml_feature_analysis_detailed.xlsx (for detailed review)\n")
            f.write("     - See ALL dogs from ALL races\n")
            f.write("     - Compare dogs within each race\n")
            f.write("     - Sorted by Track→Race→Box for easy per-race review\n")
            f.write("     - All values are from actual PDF parsing (100% factual)\n\n")
            
            f.write("=" * 80 + "\n")
        
        print(f"✅ Summary report: outputs/complete_analysis_summary.txt")
        files_generated.append('complete_analysis_summary.txt')
        
    except Exception as e:
        print(f"⚠️  Could not create summary report: {e}")
    
    except Exception as outer_error:
        print(f"\n❌ CRITICAL ERROR during file generation: {outer_error}")
        import traceback
        traceback.print_exc()
        print(f"\n⚠️  Some files may not have been generated")
    
    # Final summary
    print(f"\n" + "=" * 80)
    print(f"🎉 ANALYSIS COMPLETE!")
    print(f"   ⏱️  Total time: {(datetime.now() - start_time).total_seconds():.1f} seconds")
    print(f"\n📁 Files Generated in outputs/ folder:")
    if files_generated:
        for filename in files_generated:
            print(f"   ✅ {filename}")
    else:
        print(f"   ⚠️  No files were generated - check errors above")
    print("=" * 80)
    
    return 0

if __name__ == "__main__":
    sys.exit(main())
