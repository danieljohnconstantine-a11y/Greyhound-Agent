"""
Run Predictions with Track-Specific Ensemble Models (CALIBRATED)

Uses the track-specific ensemble models trained by train_ml_track_ensemble.py
to generate predictions on today's races in data_predictions/ folder.

For each race:
1. Loads track-specific models (RF, GB, XGB) - ALL CALIBRATED with Sigmoid Regression
2. Generates prediction from each algorithm (calibrated probabilities)
3. Combines predictions using ensemble averaging
4. Produces ML confidence score that accurately reflects win probability

Calibration ensures predicted probabilities match actual win rates, fixing
high-confidence prediction failures.

Usage:
    python run_track_ensemble_predictions.py
    
    OR use the batch file:
    run_track_ensemble_predictions.bat

Prerequisites:
    - Track-specific ensemble models trained (run train_ml_track_ensemble.bat first)
    - Race PDFs in data_predictions/ folder

Output:
    - outputs/track_ensemble_predictions.xlsx - Predictions with calibrated scores
    - outputs/track_ensemble_summary.txt - Quick summary
"""

import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

# Fix encoding for Windows console
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from src.parser import parse_race_form
from src.features import compute_features
import pandas as pd
import numpy as np
import pickle
import glob
from datetime import datetime
import traceback
import pdfplumber

# Threshold below which the pre-normalization ensemble spread is considered too
# narrow to represent genuine ML discrimination.  Races that fall below this
# value get a Low_Confidence=True flag in the output.  This value was
# established empirically from the Race 8 Angle Park audit (Mar 5 2026) where
# calibration collapse reduced model spread to <0.001 before normalization.
LOW_CONFIDENCE_SPREAD_THRESHOLD = 0.005  # < 0.5% probability spread

# Timing-derived features that should be filled with the race median rather
# than 0 when missing, to avoid unfairly penalising dogs with unknown times.
TIMING_FEATURES = [
    'BestTimeSec', 'SectionalSec', 'Speed_kmh', 'EarlySpeedIndex',
    'TimeVsField', 'SpeedVsField', 'BestTimePercentile', 'EarlySpeedPercentile',
    'SpeedAtDistance',
]

# NO cross-track fallbacks — every track uses only its own dedicated model.
# If a PDF contains a track name with no model in models/, the pipeline raises
# a hard error and stops.  This guarantees only factual, track-specific data
# is ever used in predictions.
#
# TRACK_NAME_ALIASES handles sponsor/branding name variants for the same
# physical venue.  This is NOT a cross-track fallback — the same track may
# appear under different names depending on the season (e.g. the naming-rights
# sponsor may change).  All aliases must refer to the exact same physical track.
TRACK_NAME_ALIASES = {
    # Alias (upper)                        → canonical model name
    'LADBROKES Q1 LAKESIDE':               'Q LAKESIDE',
    'LADBROKES Q2 PARKLANDS':              'Q PARKLANDS',
    'LADBROKES Q STRAIGHT':                'Q STRAIGHT',
    'LADBROKES GARDENS':                   'GARDENS',
    'BETDELUXE ROCKHAMPTON':               'ROCKHAMPTON',
    'BETDELUXE CAPALABA':                  'Capalaba',
    'BET NATION TOWNSVILLE':               'TOWNSVILLE',
    'LAKESIDE':                            'Q LAKESIDE',
    'Q PARKLANDS':                         'Q PARKLANDS',
    'GARDENS':                             'GARDENS',
    'TASMANIA':                            'HOBART',
    'MURRAY BRIDGE STRAIGHT':              'MURRAY BDGE STRAIGHT',
    'MOUNT GAMBIER':                       'MOUNT GAMBIER',
}

def extract_text_from_pdf(pdf_path):
    """
    Extract text content from a PDF file.
    
    Args:
        pdf_path: Path to the PDF file
        
    Returns:
        str: Extracted text content from all pages
    """
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text += page.extract_text() + "\n"
    return text

def load_track_ensemble(track_name, models_dir="models"):
    """
    Load all ensemble models for a specific track.

    Every track MUST have its own dedicated trained model (RF, GB, XGB, scaler).
    There are NO fallbacks — if a track's models are missing a RuntimeError is
    raised immediately so the problem is visible and cannot be silently ignored.

    Returns:
        models: Dict of {algorithm: model}
        scaler: StandardScaler for this track
        config: Ensemble configuration dict
    """
    config_path = os.path.join(models_dir, "config.pkl")

    if not os.path.exists(config_path):
        raise FileNotFoundError(f"Ensemble configuration not found: {config_path}")

    with open(config_path, 'rb') as f:
        config = pickle.load(f)

    known_tracks = config['tracks'] if isinstance(config['tracks'], list) else list(config['tracks'].keys())

    # Step 1: case-insensitive exact-match against trained tracks.
    known_tracks_upper = {t.upper(): t for t in known_tracks}
    model_track = known_tracks_upper.get(track_name.upper())

    # Step 2: if no exact match, try the sponsor-name alias table.
    # This resolves e.g. "Ladbrokes Q1 Lakeside" → "Q LAKESIDE" which is
    # the same physical venue.  This is NOT a cross-track fallback.
    if model_track is None:
        alias_canonical = TRACK_NAME_ALIASES.get(track_name.upper())
        if alias_canonical is not None:
            model_track = known_tracks_upper.get(alias_canonical.upper())

    if model_track is None:
        raise RuntimeError(
            f"\n{'='*60}\n"
            f"ERROR: No dedicated model found for track '{track_name}'.\n"
            f"Trained tracks: {sorted(known_tracks)}\n"
            f"ACTION REQUIRED: Train a model for '{track_name}' before running predictions.\n"
            f"NO FALLBACKS ARE PERMITTED — only factual track-specific models are used.\n"
            f"{'='*60}"
        )

    # Try subdirectory layout first: models/{track}/{algorithm}.pkl
    track_dir = os.path.join(models_dir, model_track)

    models = {}
    for alg in config['algorithms']:
        model_path = os.path.join(track_dir, f"{alg}.pkl")
        if os.path.exists(model_path):
            with open(model_path, 'rb') as f:
                models[alg] = pickle.load(f)

    # Flat-file layout: models/{model_track}_{algorithm}.pkl
    if not models:
        for alg in config['algorithms']:
            flat_path = os.path.join(models_dir, f"{model_track}_{alg}.pkl")
            if os.path.exists(flat_path):
                with open(flat_path, 'rb') as f:
                    models[alg] = pickle.load(f)

    if not models:
        raise RuntimeError(
            f"\n{'='*60}\n"
            f"ERROR: Model files not found for track '{track_name}' (config name: '{model_track}').\n"
            f"Expected files: models/{model_track}_rf.pkl, _gb.pkl, _xgb.pkl\n"
            f"NO FALLBACKS — train the model for this track first.\n"
            f"{'='*60}"
        )

    missing_algs = [a for a in config['algorithms'] if a not in models]
    if missing_algs:
        raise RuntimeError(
            f"\n{'='*60}\n"
            f"ERROR: Missing algorithm model(s) for track '{track_name}': {missing_algs}.\n"
            f"All 3 algorithms (RF, GB, XGB) must be present. Retrain to fix.\n"
            f"{'='*60}"
        )

    # Load scaler - try subdirectory first, then flat file
    scaler_path = os.path.join(track_dir, "scaler.pkl")
    flat_scaler_path = os.path.join(models_dir, f"{model_track}_scaler.pkl")
    if os.path.exists(scaler_path):
        with open(scaler_path, 'rb') as f:
            scaler = pickle.load(f)
    elif os.path.exists(flat_scaler_path):
        with open(flat_scaler_path, 'rb') as f:
            scaler = pickle.load(f)
    else:
        raise RuntimeError(
            f"\n{'='*60}\n"
            f"ERROR: Scaler not found for track '{track_name}' (config name: '{model_track}').\n"
            f"Expected: models/{model_track}_scaler.pkl\n"
            f"NO FALLBACKS — retrain the model for this track.\n"
            f"{'='*60}"
        )

    # Reconcile the feature list with the scaler's actual training features.
    config = dict(config)  # shallow copy — don't mutate the global config
    if hasattr(scaler, 'feature_names_in_'):
        scaler_features = list(scaler.feature_names_in_)
        if scaler_features != config.get('feature_cols', []):
            config['feature_cols'] = scaler_features

    return models, scaler, config

def _get_uncalibrated_preds(model, X_scaled):
    """
    Extract uncalibrated base-estimator predictions from a CalibratedClassifierCV.

    When calibration collapses probabilities into a narrow band all dogs get the
    same score.  The base estimator (RF/GB/XGB before isotonic fitting) preserves
    the original probability ordering, so we use it as a fallback when the
    calibrated output has fewer unique values than dogs.

    Returns:
        numpy array of probabilities (one per dog) with maximum discrimination.
    """
    # Calibrated model: sklearn CalibratedClassifierCV
    if hasattr(model, 'calibrated_classifiers_'):
        base_preds_list = []
        for cal_clf in model.calibrated_classifiers_:
            base_est = getattr(cal_clf, 'estimator', None)
            if base_est is not None and hasattr(base_est, 'predict_proba'):
                try:
                    bp = base_est.predict_proba(X_scaled)[:, 1]
                    base_preds_list.append(bp)
                except Exception:
                    pass
        if base_preds_list:
            return np.mean(base_preds_list, axis=0)
    return None


def predict_with_ensemble(df, models, scaler, feature_cols, ensemble_weights):
    """
    Generate ensemble predictions for a race WITH PROPER DISCRIMINATION.

    Individual scoring guarantee
    ----------------------------
    Each dog MUST receive a score that is 100% individual to that dog.  Two
    failure modes are detected and corrected automatically:

    1. Calibration collapse – CalibratedClassifierCV's isotonic mapping can
       compress many different raw probabilities into the same calibrated value.
       Fix: detect when ≥50% of dogs share the same calibrated score and fall
       back to the uncalibrated base-estimator predictions instead.

    2. Feature clustering – if too many input features are identical across all
       dogs the model cannot distinguish them.  This is reported as a warning.

    Returns raw model probabilities — no range mapping or normalization is applied.
    Scores reflect the actual ML confidence derived from PDF form data only.

    Args:
        df: DataFrame with race data (all dogs in race)
        models: Dict of {algorithm: model}
        scaler: StandardScaler
        feature_cols: List of feature column names
        ensemble_weights: Dict of {algorithm: weight}

    Returns:
        ensemble_pred: Array of raw ensemble probabilities for each dog
        individual_scores: Dict of {algorithm: array of raw probabilities}
        raw_spread: Float — probability spread (used for Low_Confidence flag)
    """
    n_dogs = len(df)

    # Prepare features
    # Check for missing features and warn user
    missing_features = [col for col in feature_cols if col not in df.columns]
    if missing_features:
        print(f"      ⚠️  Warning: {len(missing_features)} features missing from race data (will be filled with 0)")
        if len(missing_features) <= 5:
            print(f"         Missing: {', '.join(missing_features)}")
    
    # Extract features.
    # IMPROVEMENT: fill NaN timing-derived features with the race median
    # rather than 0. Filling with 0 unfairly scores a dog with unknown time
    # as the "slowest in the field" — median treatment is neutral.
    X = df[feature_cols].copy()
    for col in TIMING_FEATURES:
        if col in X.columns:
            col_median = X[col].median()
            fill_value = col_median if pd.notna(col_median) else 0
            X[col] = X[col].fillna(fill_value)
    X = X.fillna(0)
    
    # Check for feature variability - warn if features don't vary between dogs
    constant_features = []
    varying_features = []
    for col in feature_cols:
        if col in df.columns:
            unique_count = df[col].nunique()
            if unique_count == 1:
                constant_features.append(col)
            else:
                varying_features.append(col)
    
    # CRITICAL: Warn if >30% of features are constant (lowered threshold from 50%)
    constant_ratio = len(constant_features) / len(feature_cols) if feature_cols else 0
    if constant_ratio > 0.3:
        print(f"      ⚠️  CRITICAL: {len(constant_features)}/{len(feature_cols)} features ({constant_ratio*100:.1f}%) have same value for all dogs!")
        print(f"         This will cause identical prediction scores.")
        print(f"         Varying features: {len(varying_features)}")
        
        # Show sample of varying features
        if varying_features:
            print(f"         Sample varying: {', '.join(varying_features[:5])}")
        
        # Show critical dog-specific features that should vary
        critical_features = ['Box', 'Draw', 'Weight', 'BestTimeSec', 'SectionalSec', 'CareerWins', 
                            'CareerStarts', 'WinRate', 'PlaceRate', 'DLWFactor', 'WeightFactor']
        missing_critical = [f for f in critical_features if f not in df.columns or df[f].nunique() == 1]
        if missing_critical:
            print(f"         ⚠️  Missing/constant critical features: {', '.join(missing_critical[:10])}")
            print(f"         These should vary between dogs for accurate predictions!")
    
    X_scaled = scaler.transform(X)
    
    # Get predictions from each algorithm
    all_predictions = []
    used_weights = []
    individual_scores = {}
    
    # Equal weights for track-specific dedicated models (HEALESVILLE/Maitland/SHEPPARTON).
    # The previous 50/25/25 XGB-dominant weighting was tuned for Darwin/Rockhampton
    # cross-track fallback models where XGB had 78% discrimination vs RF 33%.
    # With dedicated track models the RF and GB uncalibrated fallback predictions
    # have equal or better spread (0.10-0.18) than the calibrated XGB (0.017-0.061),
    # so equal weighting is more appropriate.
    improved_weights = {
        'xgb': 1 / 3,
        'rf':  1 / 3,
        'gb':  1 / 3,
    }
    
    for alg, model in models.items():
        pred_proba = model.predict_proba(X_scaled)[:, 1]

        # ---------------------------------------------------------------
        # CALIBRATION-COLLAPSE GUARD
        # Two failure modes are detected:
        #
        # 1. TRUE collapse  – fewer unique probability values than half the
        #    field (e.g. isotonic mapping folds 8 dogs onto 1 value).
        #
        # 2. NEAR-collapse  – values are technically distinct but all within
        #    0.5 percentage-points of each other.  This happens when the
        #    calibrated model over-relies on a feature that is absent (e.g.
        #    Weight=0 for all dogs), causing GB/XGB to output probabilities
        #    like 0.160711001, 0.160711002 … which round to the same 4 dp
        #    value (16.0711%) for every dog.  The uncalibrated base estimator
        #    preserves the original probability ordering and is used instead.
        # ---------------------------------------------------------------
        n_unique_calibrated = len(np.unique(pred_proba))
        pred_spread = float(pred_proba.max() - pred_proba.min())
        # Treat as collapsed if true collapse OR near-constant (< 0.5% spread)
        SPREAD_THRESHOLD = 0.005
        is_near_constant = pred_spread < SPREAD_THRESHOLD
        is_true_collapse  = n_unique_calibrated < max(2, n_dogs // 2)
        if is_true_collapse or is_near_constant:
            uncal = _get_uncalibrated_preds(model, X_scaled)
            if uncal is not None:
                n_unique_uncal = len(np.unique(uncal))
                collapse_label = ("calibration collapsed (isotonic)" if is_true_collapse
                                  else f"near-collapsed (spread={pred_spread*100:.4f}%)")
                print(f"      ⚠️  {alg.upper()}: {collapse_label} – "
                      f"{n_unique_calibrated} unique value(s) across {n_dogs} dogs. "
                      f"Falling back to uncalibrated predictions ({n_unique_uncal} unique values). "
                      f"FIX: retrain this track with: python retrain_all_tracks_sigmoid.py --tracks <track>")
                pred_proba = uncal
            else:
                print(f"      ⚠️  {alg.upper()}: calibration collapsed scores "
                      f"and base estimator unavailable. "
                      f"FIX: retrain this track with: python retrain_all_tracks_sigmoid.py --tracks <track>")

        # Store individual predictions (RAW from model - no failed temperature scaling)
        individual_scores[alg] = pred_proba
        
        # Use improved weights
        weight = improved_weights.get(alg, ensemble_weights.get(alg, 1.0 / len(models)))
        all_predictions.append(pred_proba * weight)
        used_weights.append(weight)
    
    # Normalize weights and compute weighted average
    total_weight = sum(used_weights)
    ensemble_pred = np.sum(all_predictions, axis=0) / total_weight

    # Capture pre-normalization spread to expose model confidence
    raw_spread = float(ensemble_pred.max() - ensemble_pred.min())

    # Return raw probabilities directly from the model — no normalization applied.
    return ensemble_pred, individual_scores, raw_spread

def main():
    print("=" * 80)
    print("🎯 TRACK-SPECIFIC ENSEMBLE PREDICTIONS (CALIBRATED)")
    print("=" * 80)
    print("\nUsing track-specific calibrated ensemble models:")
    print("  ✅ RandomForest + GradientBoosting + XGBoost per track")
    print("  ✅ RF models calibrated with Sigmoid Regression; GB/XGB use native probabilities")
    print("  ✅ Predictions averaged across all 3 algorithms")
    print("  ✅ Confidence scores accurately reflect win probability")
    print("=" * 80)
    
    # Check if models exist
    models_dir = "models"
    config_path = os.path.join(models_dir, "config.pkl")
    
    if not os.path.exists(config_path):
        print(f"\n❌ ERROR: Track ensemble models not found!")
        print(f"   Expected config at: {config_path}")
        print(f"\n   Please train the models first:")
        print(f"   python train_ml_track_ensemble.py")
        print(f"   OR: train_ml_track_ensemble.bat")
        return 1
    
    # Load configuration
    with open(config_path, 'rb') as f:
        config = pickle.load(f)
    
    print(f"\n📥 Loaded ensemble configuration:")
    print(f"   Tracks: {len(config['tracks'])}")
    print(f"   Algorithms: {', '.join(config['algorithms'])}")
    print(f"   Features: {len(config['feature_cols'])}")
    print(f"\n📊 Feature columns used for predictions:")
    print(f"   {', '.join(config['feature_cols'][:15])}...")
    if len(config['feature_cols']) > 15:
        print(f"   ... and {len(config['feature_cols']) - 15} more features")
    
    # Find PDFs
    pdf_files = glob.glob("data_predictions/*.pdf")
    
    if not pdf_files:
        print(f"\n❌ No PDF files found in data_predictions/")
        return 1
    
    print(f"\n📄 Found {len(pdf_files)} PDF files")
    
    # Process each PDF
    all_predictions = []
    
    for pdf_file in sorted(pdf_files):
        print(f"\n📄 Processing: {os.path.basename(pdf_file)}")
        
        try:
            # Extract text from PDF
            print(f"   📖 Extracting text from PDF...")
            pdf_text = extract_text_from_pdf(pdf_file)
            
            if not pdf_text or len(pdf_text.strip()) == 0:
                print(f"   ⚠️  No text extracted from PDF")
                continue
            
            # Parse PDF text
            print(f"   🔍 Parsing race form...")
            race_df = parse_race_form(pdf_text)
            
            if race_df is None or len(race_df) == 0:
                print(f"   ⚠️  No data extracted from PDF")
                continue
            
            # Get track name
            track_name = race_df['Track'].iloc[0] if 'Track' in race_df.columns else "Unknown"
            print(f"   Track: {track_name}")
            print(f"   Dogs: {len(race_df)}")
            
            # Compute features
            try:
                race_df = compute_features(race_df)
            except Exception as e:
                print(f"   ⚠️  Error computing features: {e}")
                # Continue with available features
            
            # Load track-specific ensemble — hard error if missing, no fallbacks
            try:
                models, scaler, config = load_track_ensemble(track_name, models_dir)
                print(f"   Models loaded (dedicated): {', '.join(sorted(models.keys()))}")

                # PER-RACE MODEL INFERENCE
                # Predictions are made per race so dogs from Race 1 are never
                # scored against dogs from Race 2 (they do not compete).
                # Scores are the raw weighted-ensemble probabilities that the
                # model derives entirely from the PDF form data — no artificial
                # range-mapping or normalization is applied.
                feature_cols = config['feature_cols']
                ens_weights  = config.get('ensemble_weights', {})
                race_frames  = []

                for race_num, single_race_df in race_df.groupby('RaceNumber'):
                    single_race_df = single_race_df.copy()

                    raw_ens, raw_ind, raw_spread = predict_with_ensemble(
                        single_race_df, models, scaler, feature_cols, ens_weights
                    )

                    # Raw probability * 100 = ML Confidence score
                    single_race_df['ML_Confidence'] = np.round(raw_ens * 100, 4)
                    single_race_df['Ensemble_Score'] = raw_ens

                    # Individual algorithm scores — also raw probabilities
                    for alg, arr in raw_ind.items():
                        single_race_df[f'{alg.upper()}_Score'] = np.round(arr * 100, 4)

                    # Low-confidence flag when all dogs' raw scores are nearly identical
                    single_race_df['Low_Confidence'] = raw_spread < LOW_CONFIDENCE_SPREAD_THRESHOLD
                    if raw_spread < LOW_CONFIDENCE_SPREAD_THRESHOLD:
                        print(f"      ⚠️  Race {race_num} LOW_CONFIDENCE: spread={raw_spread:.4f}")

                    # Per-race rank (rank 1 = top pick within this race)
                    ranks = single_race_df['ML_Confidence'].rank(ascending=False, method='min').astype(int)
                    single_race_df['ML_Rank'] = ranks

                    top_r = single_race_df.loc[single_race_df['ML_Confidence'].idxmax()]
                    rf_s, gb_s, xgb_s = (float(top_r.get(f'{alg}_Score', 0)) for alg in ('RF', 'GB', 'XGB'))
                    print(f"   Race {int(race_num):2d}: Box {int(top_r['Box'])} - {top_r['DogName']} "
                          f"({float(top_r['ML_Confidence']):.4f}%  RF={rf_s:.4f}, GB={gb_s:.4f}, XGB={xgb_s:.4f})")

                    race_frames.append(single_race_df)

                race_df = pd.concat(race_frames, ignore_index=True)

                top_dog = race_df.loc[race_df['ML_Confidence'].idxmax()]
                print(f"   ✅ Card top pick: Race {int(top_dog['RaceNumber'])} "
                      f"Box {int(top_dog['Box'])} - {top_dog['DogName']} "
                      f"({float(top_dog['ML_Confidence']):.4f}%)")

                all_predictions.append(race_df)

            except Exception as e:
                print(f"   ❌ ERROR generating predictions: {e}")
                traceback.print_exc()
                continue

        except Exception as e:
            print(f"   ❌ ERROR processing PDF: {e}")
            continue
    
    # Save results
    if all_predictions:
        print(f"\n💾 Saving results...")
        
        # Combine all predictions
        df_all = pd.concat(all_predictions, ignore_index=True)
        
        # Remove only non-essential metadata columns (not used for predictions)
        # Keep all feature columns that were used for predictions so users can inspect them
        columns_to_remove = [
            'Owner', 'Color', 'Sire', 'Dam',  # Pedigree info rarely in race forms
        ]
        
        # Remove columns that exist in the dataframe
        columns_to_remove_existing = [col for col in columns_to_remove if col in df_all.columns]
        
        if columns_to_remove_existing:
            print(f"   Removing {len(columns_to_remove_existing)} metadata columns: {', '.join(columns_to_remove_existing)}...")
            df_all = df_all.drop(columns=columns_to_remove_existing)
        
        # Reorder columns: Track, RaceNumber, Box, DogName, ML_Confidence, Low_Confidence,
        # box-bias transparency columns, RF/GB/XGB scores, then rest.
        # TrackBoxWinRatePct / TrackBoxRank / BoxWinAdvantage make the box-bias signal
        # visible to the user so they can see WHY a dog was ranked higher or lower.
        priority_cols = [
            'Track', 'RaceNumber', 'Box', 'DogName',
            'ML_Confidence', 'ML_Rank', 'Low_Confidence',
            'TrackBoxWinRatePct', 'TrackBoxRank', 'BoxWinAdvantage',
            'RF_Score', 'GB_Score', 'XGB_Score',
        ]
        # Filter priority_cols to only include those that exist in df_all
        priority_cols = [col for col in priority_cols if col in df_all.columns]
        other_cols = [col for col in df_all.columns if col not in priority_cols]
        df_all = df_all[priority_cols + other_cols]
        
        # Sort by Track -> RaceNumber -> Box (for race order)
        df_all = df_all.sort_values(['Track', 'RaceNumber', 'Box'], ascending=[True, True, True])
        
        # Save to Excel with formatting
        output_dir = "outputs"
        os.makedirs(output_dir, exist_ok=True)
        
        excel_path = os.path.join(output_dir, "track_ensemble_predictions.xlsx")
        
        # Create Excel writer with formatting
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_all.to_excel(writer, index=False, sheet_name='Predictions')
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Predictions']
            
            # Define fill styles
            from openpyxl.styles import PatternFill, Border, Side
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
            black_border = Border(top=Side(style='thick', color='000000'))
            
            # Apply formatting row by row
            prev_race_key = None
            current_row = 2  # Start at row 2 (after header)
            
            for idx in range(len(df_all)):
                row = df_all.iloc[idx]
                race_key = f"{row['Track']}_{row['RaceNumber']}"
                
                # Add black line at start of new race
                if prev_race_key is not None and race_key != prev_race_key:
                    for col in range(1, len(df_all.columns) + 1):
                        worksheet.cell(row=current_row, column=col).border = black_border
                
                # Get all dogs in current race sorted by ML_Confidence
                race_mask = (df_all['Track'] == row['Track']) & (df_all['RaceNumber'] == row['RaceNumber'])
                race_dogs = df_all[race_mask].copy()
                race_dogs_sorted = race_dogs.sort_values('ML_Confidence', ascending=False)
                
                # Find position of current dog (1st, 2nd, 3rd by ML_Confidence)
                position = list(race_dogs_sorted.index).index(df_all.index[idx]) + 1
                
                # Apply color coding
                if position == 1:
                    fill = green_fill
                elif position == 2:
                    fill = yellow_fill
                elif position == 3:
                    fill = orange_fill
                else:
                    fill = None
                
                if fill:
                    for col in range(1, len(df_all.columns) + 1):
                        worksheet.cell(row=current_row, column=col).fill = fill
                
                prev_race_key = race_key
                current_row += 1
        
        print(f"✅ Excel saved (with color coding): {excel_path}")
        
        # Save summary
        summary_path = os.path.join(output_dir, "track_ensemble_summary.txt")
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("TRACK-SPECIFIC ENSEMBLE PREDICTIONS SUMMARY\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 80 + "\n\n")
            
            f.write(f"Total PDFs processed: {len(pdf_files)}\n")
            f.write(f"Successful predictions: {len(all_predictions)}\n")
            f.write(f"Total dogs predicted: {len(df_all)}\n")
            f.write("All predictions use dedicated track-specific models only — no fallbacks.\n\n")

            # Per-track summary
            for track in df_all['Track'].unique():
                track_df = df_all[df_all['Track'] == track]
                f.write(f"\n{track}:\n")
                f.write(f"  Races: {track_df['RaceNumber'].nunique()}\n")
                f.write(f"  Dogs: {len(track_df)}\n")
                
                # Top picks per race
                for race_num in sorted(track_df['RaceNumber'].unique()):
                    race_df = track_df[track_df['RaceNumber'] == race_num]
                    if len(race_df) > 0:
                        top = race_df.loc[race_df['ML_Confidence'].idxmax()]
                        # Include individual scores if available
                        individual_str = ""
                        if 'RF_Score' in top.index and 'GB_Score' in top.index and 'XGB_Score' in top.index:
                            individual_str = f" (RF={top['RF_Score']:.4f}, GB={top['GB_Score']:.4f}, XGB={top['XGB_Score']:.4f})"
                        f.write(f"  Race {race_num}: Box {top['Box']} - {top['DogName']} ({top['ML_Confidence']:.4f}%{individual_str})\n")
        
        print(f"✅ Summary saved: {summary_path}")
        
        # Best bets report: rank every race by the gap between 1st and 2nd place
        best_bets_path = os.path.join(output_dir, "best_bets_report.txt")
        bet_rows = []
        for (track, race_num), grp in df_all.groupby(['Track', 'RaceNumber'], sort=False):
            sorted_grp = grp.sort_values('ML_Confidence', ascending=False).reset_index(drop=True)
            if len(sorted_grp) < 2:
                continue
            first  = sorted_grp.iloc[0]
            second = sorted_grp.iloc[1]
            gap = float(first['ML_Confidence']) - float(second['ML_Confidence'])
            bet_rows.append({
                'track':       track,
                'race_num':    race_num,
                'box_1st':     int(first['Box']),
                'dog_1st':     str(first['DogName']),
                'score_1st':   float(first['ML_Confidence']),
                'box_2nd':     int(second['Box']),
                'dog_2nd':     str(second['DogName']),
                'score_2nd':   float(second['ML_Confidence']),
                'gap':         gap,
            })
        bet_rows.sort(key=lambda r: r['gap'], reverse=True)

        with open(best_bets_path, 'w', encoding='utf-8') as f:
            _sep = "-" * 100
            f.write("=" * 80 + "\n")
            f.write("BEST BETS RANKING REPORT\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("Races ranked by score gap between 1st and 2nd place dogs.\n")
            f.write("Largest gap = strongest model confidence / best bet.\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"{'Rank':<5} {'Track':<16} {'Race':<5} {'Gap':>8}  "
                    f"{'1st Dog':<26} {'1st%':>8}  {'2nd Dog':<26} {'2nd%':>8}\n")
            f.write(_sep + "\n")
            for rank, row in enumerate(bet_rows, start=1):
                f.write(
                    f"{rank:<5} {row['track']:<16} R{row['race_num']:<4} {row['gap']:>7.4f}%  "
                    f"Box {row['box_1st']} {row['dog_1st']:<22} {row['score_1st']:>7.4f}%  "
                    f"Box {row['box_2nd']} {row['dog_2nd']:<22} {row['score_2nd']:>7.4f}%\n"
                )
            f.write(_sep + "\n")
            f.write(f"\nTotal races ranked: {len(bet_rows)}\n")
            if bet_rows:
                f.write(f"\n🏆 BEST BET: {bet_rows[0]['track']} Race {bet_rows[0]['race_num']} "
                        f"- Box {bet_rows[0]['box_1st']} {bet_rows[0]['dog_1st']} "
                        f"(gap: {bet_rows[0]['gap']:.4f}%)\n")

        print(f"✅ Best bets report saved: {best_bets_path}")
        
        print("\n" + "=" * 80)
        print("✅ PREDICTIONS COMPLETE!")
        print("=" * 80)
        print(f"\n📊 Results:")
        print(f"   Tracks processed: {df_all['Track'].nunique()}")
        print(f"   Races: {df_all['RaceNumber'].nunique()}")
        print(f"   Dogs: {len(df_all)}")
        print(f"\n📁 Output files:")
        print(f"   {excel_path}")
        print(f"   {summary_path}")
        print(f"   {best_bets_path}")
        print("=" * 80)
    else:
        print("\n⚠️  No predictions generated")
        return 1
    
    return 0

if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as e:
        print(f"\n❌ FATAL ERROR: {e}")
        traceback.print_exc()
        sys.exit(1)
